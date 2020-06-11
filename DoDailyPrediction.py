import os
import pandas as pd
from API_Lib import get_daily_input_data
from XGBLib import create_VAR
from XGBLib import do_xgb_prediction
from XGBLib import post_process_trades
from XGBLib import create_trade_summary
from XGBLib import daily_PnL
from openpyxl import load_workbook

from dateutil.parser import parse

pd.set_option('display.max_columns', 100)
pd.set_option('display.width', 5000)
pd.set_option('max_row', 100)


root_directory = 'X:\\Production\\'
# root_directory = 'C:\\XGBoostProduction\\'


trade_handler_directory = root_directory + '\\DailyTradeFiles\\'
upload_directory = root_directory + '\\UploadFiles\\'


trade_handler_df = pd.read_excel(trade_handler_directory+'$$$TradeHandler$$$_new.xls', skiprows=9).dropna(thresh=6)
trade_handler_df = trade_handler_df.fillna('')

for row in trade_handler_df.index:
    print(trade_handler_df.iloc[[row]])

    current_isos = trade_handler_df['ISOs'][row]
    current_isos = current_isos.replace(' ','').split(',')
    date = trade_handler_df['Date'][row]
    predict_date_str_mm_dd_yyyy = date.strftime('%m_%d_%Y')
    model_type = trade_handler_df['DARTorSPREAD'][row]
    do_data_pull = trade_handler_df['DataPull'][row]
    do_prediction = trade_handler_df['Prediction'][row]
    do_postprocessing = trade_handler_df['PostProcessing_YESFiles'][row]
    do_printcharts = trade_handler_df['PrintCharts'][row]
    do_daily_PnL = trade_handler_df['DailyPnL'][row]
    do_VAR = trade_handler_df['VAR'][row]
    working_directory = trade_handler_df['WorkingDirectory'][row]
    static_directory = trade_handler_df['StaticDirectory'][row]
    model_date = trade_handler_df['ModelDate'][row]
    daily_trade_file_name = trade_handler_df['TradeVariablesFileName'][row]
    name_adder = trade_handler_df['PaperTrade?'][row]


    # Get model_date file names
    backtest_pnl_filename = model_date + '_HOURLY_BACKTEST_PnL'
    spread_files_name = model_date + '_BACKTEST_DATA_SPREAD_DART_LOCS'
    historic_var_file_name = model_date + '_VAR_DART_DICT'


    yes_dfs_dict = {}
    upload_dfs_dict = {}
    failed_locations_dict = {}

    if do_data_pull:
        daily_input_data_dict = get_daily_input_data(predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                                                     working_directory= working_directory,
                                                     static_directory=static_directory,
                                                     spread_files_name=spread_files_name)

    for iso in current_isos:

        if do_prediction:
            preds_tier1_df, preds_tier2_df, failed_locations_dict[iso] = do_xgb_prediction(predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                                                                                           iso=iso,
                                                                                           daily_trade_file_name=daily_trade_file_name,
                                                                                           working_directory= working_directory,
                                                                                           static_directory=static_directory,
                                                                                           model_type=model_type)

        if do_postprocessing:
            trades_df, yes_dfs_dict[iso], upload_dfs_dict[iso] = post_process_trades(predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                                                                                      iso=iso,
                                                                                      daily_trade_file_name=daily_trade_file_name,
                                                                                      name_adder=name_adder,
                                                                                      working_directory= working_directory,
                                                                                      static_directory=static_directory,
                                                                                      model_type=model_type)

    ## Create master upload file (contains all predictions for all model types)
    if do_postprocessing:
        # Write all upload files to same Excel file
        if model_type == 'FORCED_SPREAD':
            short_model_type = 'FSPREAD'
        else:
            short_model_type= model_type

        if model_type=='FORCED_SPREAD':
            # Open existing dart file (it will exist because forced spread model cant exist without dart model)
            orig_excel_workbook = load_workbook(upload_directory + predict_date_str_mm_dd_yyyy + '_UPLOAD_FILES_ALL_' + '.xlsx')
            excel_upload_file = pd.ExcelWriter(upload_directory + predict_date_str_mm_dd_yyyy + '_UPLOAD_FILES_ALL_' + '.xlsx', engine='openpyxl',datetime_format='mm/dd/yy')
            excel_upload_file.book = orig_excel_workbook


            # Add forced spread tabs
            for iso, upload_df in upload_dfs_dict.items():
                spread_tab_name = short_model_type + '_' + iso + '_' + name_adder
                combo_tab_name = 'COMBO' + '_' + iso + '_' + name_adder
                spread_inc_bid = min(upload_df['Bid'])
                spread_dec_bid = max (upload_df['Bid'])

                ## try to delete sheet if it already exists (it will be replaced)
                try:
                    sheet_to_remove = excel_upload_file.book.get_sheet_by_name(spread_tab_name)
                    excel_upload_file.book.remove(sheet_to_remove)
                except:
                    pass
                upload_df.to_excel(excel_upload_file, sheet_name=spread_tab_name,index=False)

                if iso not in ['ERCOT']:
                # sum/groupby dart and forced spread trades if the model is not a true spread model (ERCOT)

                    # find the most 'valid' dart upload files and use it
                    try:
                        dart_tab_name = 'DART_' + iso + '_' + ''
                        dart_df = pd.read_excel(excel_upload_file, dart_tab_name)
                    except:
                        try:
                            dart_tab_name = 'DART_' + iso + '_' + 'PprTrd'
                            dart_df = pd.read_excel(excel_upload_file, dart_tab_name)
                        except:
                            try:
                                dart_tab_name = 'DART_' + iso + '_' + 'TEST'
                                dart_df = pd.read_excel(excel_upload_file, dart_tab_name)
                            except:
                                print('No DART tab in the ISO uploader tab for ISO '+iso)
                                pass

                    dart_inc_bid = min(dart_df['Bid'])
                    dart_dec_bid = max(dart_df['Bid'])

                    combo_df = pd.concat([dart_df,upload_df],axis=0)

                    #Force numbers to strings
                    try:
                        combo_df['Node Name'] = combo_df['Node Name'].astype('int', errors='ignore')
                        combo_df['Node Name'] = combo_df['Node Name'].astype('str')
                    except:
                        pass

                    try:
                        combo_df['Node ID'] = combo_df['Node ID'].astype('int', errors='ignore')
                        combo_df['Node ID'] = combo_df['Node ID'].astype('str')
                    except:
                        pass

                    ### Sum duplicate trades
                    try:
                        #isone uses node ID
                        combo_df = combo_df.groupby(['targetdate','Node Name','Trade Type', 'Hour','Bid','Node ID']).sum()
                    except:
                        try:
                            #all other isos do not use node ID
                            combo_df = combo_df.groupby(['targetdate', 'Node ID', 'Trade Type', 'Hour', 'Bid']).sum()
                        except:
                            # all other isos do not use node ID
                            combo_df = combo_df.groupby(['targetdate', 'Node Name', 'Trade Type', 'Hour', 'Bid']).sum()

                    combo_df.reset_index(inplace=True)

                    ## Net out trades in same hour
                    combo_df.loc[combo_df['Trade Type']=='DEC','MW'] = combo_df['MW']*-1
                    combo_df.drop(columns=['Trade Type','Bid'],inplace=True)

                    ### Groupby and sum INCs and DECs in to one total MW amount per node-hour
                    try:
                        # isone uses node ID
                        combo_df = combo_df.groupby(['targetdate', 'Node Name', 'Hour', 'Node ID']).sum()
                    except:
                        try:
                            # all other isos do not use node ID
                            combo_df = combo_df.groupby(['targetdate', 'Node ID', 'Hour']).sum()
                        except:
                            combo_df = combo_df.groupby(['targetdate', 'Node Name', 'Hour']).sum()

                    combo_df.reset_index(inplace=True)

                    #### DART BID SEGMENT
                    combo_df.loc[(combo_df['MW'] < 0) & (combo_df['BidSegment'] == 1), 'Bid'] = dart_dec_bid
                    combo_df.loc[(combo_df['MW'] > 0) & (combo_df['BidSegment'] == 1), 'Bid'] = dart_inc_bid

                    #### FORCED SPREAD BID SEGMENT
                    combo_df.loc[(combo_df['MW'] < 0) & (combo_df['BidSegment'] == 2), 'Bid'] = spread_dec_bid
                    combo_df.loc[(combo_df['MW'] > 0) & (combo_df['BidSegment'] == 2), 'Bid'] = spread_inc_bid

                    #### COMBO BID SEGMENT
                    combo_df.loc[(combo_df['MW'] < 0) & (combo_df['BidSegment'] == 3), 'Bid'] = spread_dec_bid
                    combo_df.loc[(combo_df['MW'] > 0) & (combo_df['BidSegment'] == 3), 'Bid'] = spread_inc_bid

                    combo_df.loc[combo_df['MW'] < 0,'Trade Type'] = 'DEC'
                    combo_df.loc[combo_df['MW'] > 0, 'Trade Type'] = 'INC'

                    ### drop trades that netted to 0
                    combo_df = combo_df.drop(combo_df[combo_df['MW'] == 0].index)

                    combo_df['MW'] = abs(combo_df['MW'])

                    ## try to delete sheet if it already exists (it will be replaced)
                    try:
                        sheet_to_remove = excel_upload_file.book.get_sheet_by_name(combo_tab_name)
                        excel_upload_file.book.remove(sheet_to_remove)
                    except:
                        pass

                    #Format combo sheets to be in the right order
                    if iso == 'ISONE':
                        combo_df = combo_df[['targetdate', 'Node Name', 'Trade Type', 'BidSegment', 'Hour', 'MW', 'Bid', 'Node ID']].copy()
                    elif iso == 'PJM':
                        combo_df = combo_df[['targetdate', 'Node Name', 'Trade Type', 'BidSegment', 'Hour', 'MW', 'Bid']].copy()
                    else:
                        combo_df = combo_df[['targetdate', 'Node ID', 'Trade Type', 'BidSegment', 'Hour', 'MW', 'Bid']].copy()

                    combo_df.to_excel(excel_upload_file, sheet_name=combo_tab_name, index=False)


            #combine forced spread tabs with dart tabs
            excel_upload_file.book._sheets.sort(key=lambda ws: ws.title)
            excel_upload_file.save()
            excel_upload_file.close()

        else:
            try:
                #attempt to open file if it exists. First load the existing workbook and then save a new workbook with the same name, setting it equal to the old workbook
                orig_excel_workbook = load_workbook(upload_directory + predict_date_str_mm_dd_yyyy + '_UPLOAD_FILES_ALL_' + '.xlsx')
                excel_upload_file = pd.ExcelWriter(upload_directory + predict_date_str_mm_dd_yyyy + '_UPLOAD_FILES_ALL_' + '.xlsx', engine='openpyxl',datetime_format='mm/dd/yy')
                excel_upload_file.book = orig_excel_workbook

                #Delete the sheet if it already exists from a previous model run
                try:
                    sheet_to_remove = excel_upload_file.book.get_sheet_by_name(short_model_type + '_' + iso + '_' + name_adder)
                    excel_upload_file.book.remove(sheet_to_remove)
                except:
                    pass

                for iso, upload_df in upload_dfs_dict.items():
                    upload_df.to_excel(excel_upload_file, sheet_name=short_model_type + '_' + iso + '_' + name_adder, index=False)

                excel_upload_file.book._sheets.sort(key=lambda ws: ws.title)
                excel_upload_file.save()
                excel_upload_file.close()
            except:
                #Create file since it doesnt exist
                excel_upload_file = pd.ExcelWriter(upload_directory + predict_date_str_mm_dd_yyyy + '_UPLOAD_FILES_ALL_' + '.xlsx', engine='openpyxl',datetime_format='mm/dd/yy')
                for iso, upload_df in upload_dfs_dict.items():
                    upload_df.to_excel(excel_upload_file, sheet_name=short_model_type+'_'+iso+'_'+name_adder, index=False)

                excel_upload_file.book._sheets.sort(key=lambda ws: ws.title)
                excel_upload_file.save()
                excel_upload_file.close()


    if do_printcharts and do_postprocessing:
        create_trade_summary(predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                             isos=current_isos,
                             do_printcharts=do_printcharts,
                             name_adder=name_adder,
                             working_directory= working_directory,
                             static_directory=static_directory,
                             model_type=model_type)

    if do_VAR:
        create_VAR(historic_var_file_name=historic_var_file_name,
                   working_directory=working_directory,
                   static_directory=static_directory,
                   predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                   name_adder=name_adder)

    if do_daily_PnL:
        daily_PnL(predict_date_str_mm_dd_yyyy=predict_date_str_mm_dd_yyyy,
                  isos=current_isos,
                  name_adder=name_adder,
                  working_directory=working_directory,
                  static_directory=static_directory,
                  do_printcharts=do_printcharts,
                  backtest_pnl_filename=backtest_pnl_filename,
                  model_type=model_type,
                  spread_files_name=spread_files_name)





print('')
print('**********************************************************************************')
print('')
print('$$$$$$$$$$$$$$$$$$$$$$$$$$$$$ submit trades get rich $$$$$$$$$$$$$$$$$$$$$$$$$$$$$')
print('')
